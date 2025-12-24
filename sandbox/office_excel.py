from .executor import execute_python_code
from .excel_operations import (
    add_column_operation,
    filter_operation,
    update_cell_operation,
    add_row_operation,
    delete_rows_operation,
    create_summary_operation,
)
from .excel_operations.formatting import get_copy_formatting_code, get_smart_format_code

# Common helper function code to be injected
FIND_EXCEL_FILE_FUNC = '''
def find_excel_file(target_name=None):
    files = os.listdir('/app/data')
    if target_name:
        if target_name in files: return target_name
        for ext in ['.xlsx', '.xls']:
            if f"{target_name}{ext}" in files: return f"{target_name}{ext}"
    
    excel_files = [f for f in files if f.endswith(('.xlsx', '.xls'))]
    # Khi EDIT, ưu tiên file đã edit trước đó để có thể sửa tiếp (chaining)
    # Nhưng nếu edit lần đầu thì lấy file gốc
    edited_files = [f for f in excel_files if f.endswith('_edited.xlsx')]
    if edited_files: return edited_files[0]
    if excel_files: return excel_files[0]
    return None
'''


def get_excel_sheets(session_id: str, filename: str = None) -> str:
    """Liệt kê danh sách các sheet trong file Excel.
    
    Args:
        session_id: ID của session để thực thi code
        filename: Tên file cụ thể (tùy chọn)
    """
    code = f'''
import pandas as pd
import os

{FIND_EXCEL_FILE_FUNC}

filename = find_excel_file({repr(filename)})
if not filename:
    print("Can't find Excel file!")
    exit(1)

try:
    xl = pd.ExcelFile(f'/app/data/{{filename}}')
    print(f"Sheets in file '{{filename}}':")
    for sheet in xl.sheet_names:
        print(f"- {{sheet}}")
except Exception as e:
    if "not a zip file" in str(e).lower() or "BadZipFile" in str(type(e)):
        print(f"Error: The file '{{filename}}' is corrupted or not a valid Excel (.xlsx) file.")
        print("This often happens if a previous operation timed out while saving.")
    else:
        print(f"Error: {{e}}")
    exit(1)
'''
    return execute_python_code(code, session_id)


def read_excel_content(session_id: str, filename: str = None, sheet_name: str = None, max_rows: int = 10) -> str:
    """Đọc nội dung Excel.
    
    Args:
        session_id: ID của session để thực thi code
        filename: Tên file cụ thể (tùy chọn)
        sheet_name: Tên sheet cần đọc (Nên gọi get_excel_sheets trước để biết tên sheet)
        max_rows: Số dòng tối đa hiển thị
    """
    code = f'''
import pandas as pd
import os

{FIND_EXCEL_FILE_FUNC}

filename = find_excel_file({repr(filename)})
if not filename:
    print("Can't find Excel file!")
    exit(1)

try:
    # Preview: Only read first few rows to save memory/time
    df = pd.read_excel(f'/app/data/{{filename}}', sheet_name={repr(sheet_name)}, nrows={max_rows})
    if isinstance(df, dict):
        first_sheet = list(df.keys())[0]
        print(f"Report: File has multiple sheets. Displaying first sheet '{{first_sheet}}'")
        df = df[first_sheet]
    print(df.to_string())
except Exception as e:
    if "not a zip file" in str(e).lower() or "BadZipFile" in str(type(e)):
        print(f"Error: The file '{{filename}}' is corrupted or not a valid Excel (.xlsx) file.")
        print("This often happens if the previous operation timed out while saving.")
    else:
        print(f"Error: {{e}}")
    exit(1)
'''
    return execute_python_code(code, session_id)

def edit_excel_document(session_id: str, operations: list, filename: str = None, sheet_name: str = None) -> str:
    """Edit Excel document with structured operations.
    You should read the file first to get the content by using read_excel_content operation.
    Then you can edit the file by using other operations.
    
    Args:
        session_id: ID of the session to execute code
        operations: List of operations to perform. For 'custom_code', avoid 'to_excel' or 'ExcelWriter' 
                    to preserve formatting; use 'ws.cell()' and 'wb.save()' instead.
                    Tip: Use 'copy_cell_formatting(src, dst)' or 'apply_smart_format(cell, val)' 
                    to maintain consistency for new data.
        filename: Name of the file to edit (optional)
        sheet_name: Name of the sheet to edit (optional, default to first sheet)
    """
    # Map operation types to their handler functions
    operation_handlers = {
        'add_column': add_column_operation,
        'filter': filter_operation,
        'update_cell': update_cell_operation,
        'add_row': add_row_operation,
        'delete_rows': delete_rows_operation,
        'create_summary': create_summary_operation,
    }
    
    # Build the operation code
    operations_code = []
    for op in operations:
        op_type = op.get('type')
        
        if op_type == 'custom_code':
            # Handle custom code directly
            custom_code = op.get('code')
            
            # VALIDATION: Phát hiện pattern nguy hiểm làm mất định dạng
            bad = {'ExcelWriter': 'LOST FORMAT', 'writer.book': 'CAUSE ERROR', 'to_excel': 'LOST FORMAT'}
            warns = [f"{p}: {m}" for p, m in bad.items() if p in custom_code]
            msg = f'print("⚠️ WARNING: " + {repr(", ".join(warns))} + "\\n💡 Use ws.cell() and wb.save() to keep format.")' if warns else ""
            
            operations_code.append(f'''
# Custom code operation
{msg}
try:
    exec({repr(custom_code)}, globals())
    print("- Done custom code.")
except Exception as e:
    print(f"- Error: {{e}}")
    import traceback
    traceback.print_exc()
''')
        elif op_type in operation_handlers:
            # Use the appropriate handler
            operations_code.append(operation_handlers[op_type](op))
        else:
            operations_code.append(f'''
print(f"- Warning: Operation type '{op_type}' is not supported")
''')
    
    # Combine all operations into final code
    code = f'''
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
import json

{FIND_EXCEL_FILE_FUNC}
{get_copy_formatting_code()}
{get_smart_format_code()}

filename = find_excel_file({repr(filename)})
if not filename:
    print("Error: File Excel not found!")
    exit(1)

file_path = f'/app/data/{{filename}}'
print(f"Processing file: {{filename}}")

# Read Excel file with all formats
# data_only=False: keep formulas instead of values
# keep_vba=True: keep VBA macros if any
try:
    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
except Exception as e:
    # If can't open using openpyxl (maybe .xls or corrupted)
    if "not a zip file" in str(e).lower() or "BadZipFile" in str(type(e)):
        print(f"Error: The file '{filename}' is corrupted or not a valid Excel (.xlsx) file.")
        print("This often happens if a previous operation timed out while saving.")
        print("Tip: You can try downloading the original file again to start fresh.")
    else:
        print(f"Error: Can't open file using openpyxl engine. {{e}}")
    exit(1)

target_sheet = {repr(sheet_name)}
if target_sheet and target_sheet in wb.sheetnames:
    ws = wb[target_sheet]
else:
    target_sheet = wb.sheetnames[0]
    ws = wb[target_sheet]

print(f"Processing sheet: {{target_sheet}}")

# Read using pandas for easier data processing
df = pd.read_excel(file_path, sheet_name=target_sheet)
df.columns = [c.strip() for c in df.columns]

{chr(10).join(operations_code)}

# Save edited file with all formats
base, ext = os.path.splitext(filename)
# Ensure ext is always .xlsx if file origin doesn't have ext or is .xls (convert to .xlsx)
if not ext or ext.lower() == '.xls':
    ext = '.xlsx'

if base.endswith('_edited'):
    output_filename = f"{{base}}{{ext}}"
else:
    output_filename = f"{{base}}_edited{{ext}}"

save_path = f'/app/data/{{output_filename}}'

# Ensure workbook properties are kept
# This helps keep metadata like author, created date, etc.
try:
    # Save with all options to keep format
    wb.save(save_path)
    print(f"\\nDone saving edited file: {{output_filename}}")
    print("All original formats have been kept.")
except Exception as e:
    print(f"Error when saving file: {{e}}")
    # Try saving with another way if error
    try:
        wb.save(save_path)
        print(f"\\nDone saving file (fallback): {{output_filename}}")
    except Exception as e2:
        print(f"Can't save file: {{e2}}")
        exit(1)
'''
    return execute_python_code(code, session_id)

