"""Formatting utilities for Excel operations."""


def get_copy_formatting_code() -> str:
    """Trả về code Python để sao chép định dạng từ cell nguồn sang cell đích.
    
    Returns:
        Python code dạng string để sao chép định dạng
    """
    return '''
def copy_cell_formatting(source_cell, target_cell):
    """Sao chép định dạng từ source_cell sang target_cell một cách triệt để.
    """
    from copy import copy
    
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)
    
    # Sao chép độ rộng cột nếu là cell ở các dòng đầu tiên
    if source_cell.row == 1:
        from openpyxl.utils import get_column_letter
        source_col_letter = get_column_letter(source_cell.column)
        target_col_letter = get_column_letter(target_cell.column)
        if hasattr(source_cell.parent, 'column_dimensions'):
            if source_col_letter in source_cell.parent.column_dimensions:
                source_dim = source_cell.parent.column_dimensions[source_col_letter]
                target_cell.parent.column_dimensions[target_col_letter].width = source_dim.width
                target_cell.parent.column_dimensions[target_col_letter].hidden = source_dim.hidden

    # Sao chép độ cao dòng nếu là cell ở cột đầu tiên
    if source_cell.column == 1:
        source_row = source_cell.row
        target_row = target_cell.row
        if hasattr(source_cell.parent, 'row_dimensions'):
            if source_row in source_cell.parent.row_dimensions:
                source_dim = source_cell.parent.row_dimensions[source_row]
                target_cell.parent.row_dimensions[target_row].height = source_dim.height
                target_cell.parent.row_dimensions[target_row].hidden = source_dim.hidden
'''

def get_smart_format_code() -> str:
    """Trả về code Python để tự động áp dụng định dạng số thông minh.
    
    Returns:
        Python code dạng string
    """
    return '''
def apply_smart_format(target_cell, value, col_name="", source_cell=None):
    """Áp dụng định dạng số cho cell. 
    Ưu tiên sao chép từ source_cell. Nếu không có, tự động nhận diện dựa trên giá trị và tên cột.
    """
    if not isinstance(value, (int, float)):
        return
        
    # 1. Ưu tiên cao nhất: Sao chép định dạng từ cell mẫu (source_cell)
    if source_cell and source_cell.number_format and source_cell.number_format != 'General':
        target_cell.number_format = source_cell.number_format
        return

    # 2. Nếu cell đích đã có định dạng sẵn (không phải General), giữ nguyên
    if target_cell.number_format and target_cell.number_format != 'General':
        return
        
    # 3. Fallback: Tự động nhận diện (nếu không có mẫu để copy)
    col_name_lower = str(col_name).lower()
    is_percentage = any(kw in col_name_lower for kw in ['%', 'ratio', 'rate', 'margin', 'tỉ lệ', 'phần trăm'])
    
    if is_percentage:
        target_cell.number_format = '0.00%'
    elif isinstance(value, float):
        is_currency = any(kw in col_name_lower for kw in ['price', 'cost', 'amount', 'revenue', 'giá', 'tiền', 'lương', 'vốn'])
        if is_currency or abs(value) >= 100:
            target_cell.number_format = '#,##0.00'
        else:
            target_cell.number_format = '0.00'
    else:
        if abs(value) >= 1000:
            target_cell.number_format = '#,##0'
        else:
            target_cell.number_format = '0'
'''

