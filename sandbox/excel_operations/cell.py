"""Cell operations for Excel documents."""

from .formatting import get_copy_formatting_code, get_smart_format_code


def update_cell_operation(op: dict) -> str:
    """Tạo code để cập nhật cell trong Excel document với định dạng giữ nguyên hoặc tối ưu.
    
    Args:
        op: Dictionary chứa 'row', 'column', và 'value'
        
    Returns:
        Python code để thực hiện cập nhật cell
    """
    row = op.get('row')
    col_name = op.get('column')
    value = op.get('value')
    
    code = f'''
# Update cell operation with formatting
{get_copy_formatting_code()}
{get_smart_format_code()}

row = {row}
col_name = {repr(col_name)}
value = {repr(value)}

# Tìm index cột nếu là tên
if isinstance(col_name, str) and col_name in df.columns:
    col_idx = list(df.columns).index(col_name) + 1
else:
    # Thử parse chữ cái cột (A, B, C...)
    try:
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(str(col_name))
    except:
        col_idx = None

if col_idx:
    target_cell = ws.cell(row=row, column=col_idx)
    
    # Tìm cell mẫu cùng cột (thường là dòng tiêu đề hoặc dòng dữ liệu khác) để lấy định dạng
    source_row = 2 if row != 2 else (3 if ws.max_row >= 3 else 1)
    source_cell = ws.cell(row=source_row, column=col_idx) if source_row <= ws.max_row else None
    
    # Nếu cell chưa có style, thử sao chép định dạng đầy đủ (font, border, fill...)
    if not target_cell.has_style and source_cell:
        copy_cell_formatting(source_cell, target_cell)
    
    # Cập nhật giá trị
    target_cell.value = value
    
    # Tinh chỉnh định dạng số (Ưu tiên copy đơn vị từ source_cell)
    apply_smart_format(target_cell, value, col_name, source_cell)

    print(f"- Đã cập nhật ô tại dòng {{row}}, cột {{col_name}} thành '{{value}}' với đơn vị được bảo toàn")
else:
    print(f"- Không tìm thấy cột '{{col_name}}' để cập nhật")
'''
    return code
