from .formatting import get_copy_formatting_code, get_smart_format_code


def add_column_operation(op: dict) -> str:
    """Tạo code để thêm cột vào Excel document với định dạng giống các cột khác.
    
    Args:
        op: Dictionary chứa 'name' và 'formula' của cột mới
    """
    col_name = op.get('name')
    formula = op.get('formula')
    
    code = f'''
# Add column operation with formatting
{get_copy_formatting_code()}
{get_smart_format_code()}

col_name = {repr(col_name)}
formula = {repr(formula)}
try:
    df[col_name] = df.eval(formula)
    new_col_idx = ws.max_column + 1
    
    # Tìm cột mẫu để sao chép định dạng (lấy cột cuối cùng trước cột mới)
    source_col_idx = ws.max_column
    
    # Thêm header
    new_header = ws.cell(row=1, column=new_col_idx, value=col_name)
    if source_col_idx > 0:
        source_header = ws.cell(row=1, column=source_col_idx)
        copy_cell_formatting(source_header, new_header)
    
    # Thêm dữ liệu và sao chép định dạng
    for i, val in enumerate(df[col_name], start=2):
        new_cell = ws.cell(row=i, column=new_col_idx, value=val)
        
        if source_col_idx > 0:
            source_cell = ws.cell(row=i, column=source_col_idx)
            copy_cell_formatting(source_cell, new_cell)
            
            # Tinh chỉnh định dạng số và đơn vị (Tiền tệ, %, VND...)
            apply_smart_format(new_cell, val, col_name, source_cell)
    
    # Cập nhật AutoFilter để bao phủ cả cột mới
    from openpyxl.utils import get_column_letter
    full_range = f"A1:{{get_column_letter(new_col_idx)}}{{ws.max_row}}"
    ws.auto_filter.ref = full_range
        
    print(f"- Added column '{{col_name}}' with formula '{{formula}}' with professional formatting and updated filter")
except Exception as e:
    print(f"- Error adding column '{{col_name}}': {{e}}")
    import traceback
    traceback.print_exc()
'''
    return code
